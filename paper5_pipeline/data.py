"""Data loading, recoding, MICE imputation, encoding, and equity strata.

Produces a `Cohort` object holding the encoded node-feature matrix X (the GAT
inputs), the 8 target scales Y, a binary label for the classification head, the
analytic raw frame, and the three equity-stratum vectors.
"""
from __future__ import annotations

from dataclasses import dataclass

import numpy as np
import openpyxl
import pandas as pd

from . import config as C


# ------------------------------------------------------------------ recoders #
def _recode_contraceptive(v):
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return np.nan
    v = int(v)
    if v == 0:
        return "none"
    if v in (1, 2, 4, 6):
        return "hormonal_systemic"
    if v in (3, 5):
        return "hormonal_local"
    if v == 7:
        return "mechanical"
    return "other"  # 888 / unknown


def _recode_smoking(v):
    # 0 never, 1 current, 2 ex -> ordinal never(0) < ex(1) < current(2)
    return {0: 0, 2: 1, 1: 2}.get(v, np.nan)


def _read_excel(path) -> pd.DataFrame:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = [str(h) for h in next(it)]
    rows = [list(r) for r in it]
    wb.close()
    return pd.DataFrame(rows, columns=header)


@dataclass
class Cohort:
    X: np.ndarray            # (n, d) standardised encoded node features
    feat_names: list         # length-d encoded feature names
    Y: pd.DataFrame          # (n, 8) target scales 0-100 (NaN where unobserved)
    y_bin: np.ndarray        # (n,) binary label for BINARY_TARGET (NaN where unobserved)
    raw: pd.DataFrame        # analytic raw frame (n rows)
    strata: pd.DataFrame     # education / partnership / diagnosis stratum columns
    raw_features: pd.DataFrame  # pre-encoding imputed feature frame
    feat_missing: dict       # pre-imputation missingness % per feature (cohort)


def _map_features(raw: pd.DataFrame) -> pd.DataFrame:
    d = pd.DataFrame(index=raw.index)
    num = lambda c: pd.to_numeric(raw[c], errors="coerce")
    # --- sociodemographic / reproductive ---
    d["education"] = num("education")
    d["marital_status"] = num("marital_status")
    d["age"] = num("age")
    bmi = num("bmi")
    lo, hi = bmi.quantile([0.01, 0.99])
    d["bmi"] = bmi.clip(lo, hi)
    # raw menarche here; out-of-range winsorisation happens AFTER the
    # complete-case mask so plausibility-editing does not shrink the cohort.
    d["menarche"] = num("menstruation_firsttime_age")
    d["pregnancies"] = num("pregnancy_number")
    d["menopausal"] = num("menopause_yn")
    d["contraceptive"] = raw["contraceptive_kind"].map(_recode_contraceptive)
    d["alcohol"] = num("alcohol")
    d["smoking"] = raw["smokingstatus"].map(_recode_smoking)
    d["bust"] = num("bust")
    d["cup"] = num("cupsize")
    # --- clinical (low missingness) ---
    d["diagnosis"] = num("diagnosis")
    d["pre_op"] = num("pre_op")
    d["breastcancer_first"] = num("breastcancer_first")
    d["cancer_breast"] = num("cancer_breast")
    for c in C.COMORB:
        d[c] = num(c)
    return d


def _mice_impute(df: pd.DataFrame, seed: int) -> pd.DataFrame:
    """MICE on numeric-coded features. Nominal strings are factorised first."""
    from sklearn.experimental import enable_iterative_imputer  # noqa: F401
    from sklearn.impute import IterativeImputer

    work = df.copy()
    cat_maps = {}
    for col in ["contraceptive"]:
        codes, uniques = pd.factorize(work[col])
        codes = codes.astype(float)
        codes[codes < 0] = np.nan
        work[col] = codes
        cat_maps[col] = list(uniques)
    imp = IterativeImputer(max_iter=10, random_state=seed, sample_posterior=False)
    arr = imp.fit_transform(work.values)
    out = pd.DataFrame(arr, columns=work.columns, index=work.index)
    # round discrete columns back to valid codes
    discrete = (["education", "marital_status", "pregnancies", "menopausal",
                 "alcohol", "smoking", "bust", "cup", "diagnosis", "pre_op",
                 "breastcancer_first", "cancer_breast"] + C.COMORB)
    for col in discrete:
        if col in out:
            out[col] = out[col].round().clip(df[col].min(), df[col].max())
    out["contraceptive"] = out["contraceptive"].round().clip(0, len(cat_maps["contraceptive"]) - 1)
    out["contraceptive"] = out["contraceptive"].map(lambda i: cat_maps["contraceptive"][int(i)])
    return out


def _encode(feat: pd.DataFrame):
    """One-hot nominal, keep continuous, pass binaries; return matrix + names."""
    cols, names = [], []
    for c in C.CONTINUOUS:
        cols.append(feat[c].astype(float).values.reshape(-1, 1))
        names.append(c)
    # education 1/2/3
    for lvl in [1, 2, 3]:
        cols.append((feat["education"].round() == lvl).astype(float).values.reshape(-1, 1))
        names.append(f"education_{C.EDU_STRATA[lvl]}")
    # marital 0/1/2/3
    for lvl, lab in [(0, "none"), (1, "partner"), (2, "div_sep"), (3, "widowed")]:
        cols.append((feat["marital_status"].round() == lvl).astype(float).values.reshape(-1, 1))
        names.append(f"marital_{lab}")
    # contraceptive one-hot
    for lab in ["none", "hormonal_systemic", "hormonal_local", "mechanical", "other"]:
        cols.append((feat["contraceptive"] == lab).astype(float).values.reshape(-1, 1))
        names.append(f"contraceptive_{lab}")
    # diagnosis 1/2/3/4 one-hot
    for lvl, lab in [(1, "breast_cancer"), (2, "dcis"), (3, "fibroadenoma"), (4, "other_benign")]:
        cols.append((feat["diagnosis"].round() == lvl).astype(float).values.reshape(-1, 1))
        names.append(f"diagnosis_{lab}")
    # binary flags (menopausal + clinical + comorbidities)
    for c in C.BINARY:
        cols.append((feat[c].round() == 1).astype(float).values.reshape(-1, 1))
        names.append(c if c.startswith("comorb_") else f"{c}_yes")
    X = np.hstack(cols)
    return X, names


def load_cohort(seed: int = C.SEED) -> Cohort:
    raw_all = _read_excel(C.RAW_XLSX)
    for c in raw_all.columns:
        if c not in ("contraceptive_kind", "smokingstatus"):
            pass
    feat_all = _map_features(raw_all)

    # analytic cohort = complete-case across the sociodemographic/reproductive
    # block (matches n=1434); clinical missingness is MICE-imputed within cohort.
    cc = feat_all[C.COHORT_COMPLETE_CASE].notna().all(axis=1)
    raw = raw_all[cc].reset_index(drop=True)
    feat = feat_all[cc].reset_index(drop=True)

    # winsorise implausible menarche to NaN, then MICE-impute within the cohort
    feat["menarche"] = feat["menarche"].where(
        (feat["menarche"] >= 8) & (feat["menarche"] <= 18))
    feat_missing = {c: float(feat[c].isna().mean() * 100) for c in feat.columns}
    if feat.isna().any().any():
        feat = _mice_impute(feat, seed)

    X_raw, names = _encode(feat)
    # standardise continuous columns only (one-hot left as 0/1)
    X = X_raw.astype(float).copy()
    cont_idx = [names.index(c) for c in C.CONTINUOUS]
    mu = X[:, cont_idx].mean(0)
    sd = X[:, cont_idx].std(0)
    sd[sd == 0] = 1.0
    X[:, cont_idx] = (X[:, cont_idx] - mu) / sd

    # targets
    Y = pd.DataFrame({t: pd.to_numeric(raw[t], errors="coerce") for t in C.TARGETS})
    yb = Y[C.BINARY_TARGET].values
    y_bin = np.where(np.isnan(yb), np.nan, (yb < C.BINARY_CUTOFF).astype(float))

    # equity strata
    strata = pd.DataFrame(index=range(len(raw)))
    strata["education"] = feat["education"].round().map(C.EDU_STRATA).values
    strata["partnership"] = [C.partnership(int(m)) for m in feat["marital_status"].round()]
    diag = pd.to_numeric(raw["diagnosis"], errors="coerce")
    strata["diagnosis"] = np.where(diag.isin(list(C.DIAG_MALIGNANT)), "malignant", "benign")

    return Cohort(X=X, feat_names=names, Y=Y, y_bin=y_bin, raw=raw,
                  strata=strata, raw_features=feat, feat_missing=feat_missing)
